"""
converters/pdf.py — PDF as a real input, not just an OCR target.

Three-tier strategy (each library is an optional dep; clear error if missing):

  pymupdf (fitz)  — fast, accurate text + layout extraction for digital PDFs.
                    Also renders page images and exports native HTML.
  pdfplumber      — table extraction into structured data (CSV / XLSX / JSON).
  pdf2docx        — layout-preserving PDF → DOCX conversion.

Fallback chain for pdf → txt:
  1. pymupdf      (digital text, fastest, highest quality)
  2. pdfplumber   (good fallback for some layouts pymupdf misses)
  3. pytesseract  (OCR — only for scanned / image-only PDFs)

All routes respect --page N (1-indexed) to extract a single page.
"""

from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Optional

from ..registry import ConversionResult, OptionSpec, register

# ── shared options ─────────────────────────────────────────────────────────

_PAGE_OPT = OptionSpec(
    "page", ("-p", "--page"),
    "Extract only this page number (1-indexed). Default: all pages.",
    type=int,
)
_DPI_OPT = OptionSpec(
    "dpi", ("--dpi",),
    "Render DPI for image export (pdf → png/jpg). Default: 150.",
    default=150, type=int,
)
_PASSWORD_OPT = OptionSpec(
    "password", ("--password",),
    "Password for encrypted PDFs.",
)

_TEXT_OPTIONS  = [_PAGE_OPT, _PASSWORD_OPT]
_IMAGE_OPTIONS = [_PAGE_OPT, _DPI_OPT, _PASSWORD_OPT]
_TABLE_OPTIONS = [_PAGE_OPT, _PASSWORD_OPT]


# ── helpers ────────────────────────────────────────────────────────────────

def _require_fitz():
    try:
        import fitz  # noqa: F401 (pymupdf)
        return fitz
    except ImportError:
        raise RuntimeError(
            "pymupdf is required for PDF text/image extraction.\n"
            "Install it with: pip install pymupdf\n"
            "Or: pip install morphconv[pdf]"
        )


def _require_pdfplumber():
    try:
        import pdfplumber  # noqa: F401
        return pdfplumber
    except ImportError:
        raise RuntimeError(
            "pdfplumber is required for PDF table extraction.\n"
            "Install it with: pip install pdfplumber\n"
            "Or: pip install morphconv[pdf]"
        )


def _require_pdf2docx():
    try:
        from pdf2docx import Converter as _Cv  # noqa: F401
        return _Cv
    except ImportError:
        raise RuntimeError(
            "pdf2docx is required for layout-preserving PDF → DOCX conversion.\n"
            "Install it with: pip install pdf2docx\n"
            "Or: pip install morphconv[pdf]"
        )


def _open_fitz(input_path: Path, password: Optional[str] = None):
    fitz = _require_fitz()
    doc = fitz.open(str(input_path))
    if doc.is_encrypted:
        if not password:
            raise RuntimeError(
                "This PDF is encrypted. Provide the password with --password."
            )
        if not doc.authenticate(password):
            raise RuntimeError("Incorrect PDF password.")
    return doc


def _page_range(doc, page: Optional[int]) -> range:
    """Convert a 1-indexed user page number to a range over 0-indexed pages."""
    if page is not None:
        idx = page - 1
        if idx < 0 or idx >= len(doc):
            raise RuntimeError(
                f"Page {page} out of range — this PDF has {len(doc)} page(s)."
            )
        return range(idx, idx + 1)
    return range(len(doc))


# ── pdf → txt ─────────────────────────────────────────────────────────────

def _pdf_to_txt(
    input_path: Path,
    output_path: Path,
    *,
    page: Optional[int] = None,
    password: Optional[str] = None,
    **_options,
) -> ConversionResult:
    # Tier 1: pymupdf
    try:
        doc = _open_fitz(input_path, password)
        pages = _page_range(doc, page)
        chunks: list[str] = []
        for i in pages:
            text = doc[i].get_text("text")
            chunks.append(text)
        doc.close()
        full_text = "\n\f\n".join(chunks)  # form-feed between pages
        if full_text.strip():
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(full_text, encoding="utf-8")
            return ConversionResult(output=output_path, pages=len(chunks))
    except RuntimeError as e:
        if "pymupdf is required" not in str(e):
            raise

    # Tier 2: pdfplumber
    try:
        pdfplumber = _require_pdfplumber()
        chunks = []
        with pdfplumber.open(str(input_path), password=password or "") as pdf:
            pages_list = [pdf.pages[page - 1]] if page else pdf.pages
            for pg in pages_list:
                text = pg.extract_text() or ""
                chunks.append(text)
        full_text = "\n\f\n".join(chunks)
        if full_text.strip():
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(full_text, encoding="utf-8")
            return ConversionResult(output=output_path, pages=len(chunks))
    except RuntimeError as e:
        if "pdfplumber is required" not in str(e):
            raise

    # Tier 3: OCR fallback via pytesseract (scanned PDFs)
    try:
        import pytesseract
        from PIL import Image as PILImage
    except ImportError:
        raise RuntimeError(
            "No PDF text extraction library is available.\n"
            "Install pymupdf: pip install pymupdf\n"
            "Or install the full PDF stack: pip install morphconv[pdf]"
        )

    try:
        fitz = _require_fitz()
        doc = _open_fitz(input_path, password)
        pages_range = _page_range(doc, page)
        chunks = []
        for i in pages_range:
            pix = doc[i].get_pixmap(dpi=200)
            img = PILImage.frombytes("RGB", [pix.width, pix.height], pix.samples)
            text = pytesseract.image_to_string(img)
            chunks.append(text)
        doc.close()
    except Exception:
        raise RuntimeError(
            "PDF text extraction failed. The PDF may be scanned/image-only.\n"
            "Install pymupdf + pytesseract for OCR fallback:\n"
            "  pip install morphconv[pdf]\n"
            "  (and ensure Tesseract is on your PATH)"
        )

    full_text = "\n\f\n".join(chunks)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(full_text, encoding="utf-8")
    return ConversionResult(output=output_path, pages=len(chunks))


# ── pdf → html ────────────────────────────────────────────────────────────

def _pdf_to_html(
    input_path: Path,
    output_path: Path,
    *,
    page: Optional[int] = None,
    password: Optional[str] = None,
    **_options,
) -> ConversionResult:
    fitz = _require_fitz()
    doc = _open_fitz(input_path, password)
    pages = _page_range(doc, page)

    html_parts: list[str] = [
        "<!DOCTYPE html><html><head><meta charset='utf-8'></head><body>"
    ]
    for i in pages:
        html_parts.append(f"<!-- page {i + 1} -->")
        html_parts.append(doc[i].get_text("html"))
        html_parts.append("<hr>")
    html_parts.append("</body></html>")
    doc.close()

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(html_parts), encoding="utf-8")
    return ConversionResult(output=output_path, pages=len(pages))


# ── pdf → md ──────────────────────────────────────────────────────────────

def _pdf_to_md(
    input_path: Path,
    output_path: Path,
    *,
    page: Optional[int] = None,
    password: Optional[str] = None,
    **_options,
) -> ConversionResult:
    """
    Extracts PDF text as Markdown. Uses pymupdf's block-level layout info to
    infer headings from font size (largest sizes become h1/h2/h3) and bold
    spans become **bold**.
    """
    fitz = _require_fitz()
    doc = _open_fitz(input_path, password)
    pages = _page_range(doc, page)

    # Collect all font sizes to build heading thresholds
    all_sizes: list[float] = []
    for i in pages:
        blocks = doc[i].get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]
        for block in blocks:
            if block["type"] != 0:
                continue
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    sz = span.get("size", 0)
                    if sz > 0:
                        all_sizes.append(sz)

    all_sizes_sorted = sorted(set(all_sizes), reverse=True)
    # Only treat a size as a heading if it's clearly larger than body text
    body_size = all_sizes_sorted[-1] if all_sizes_sorted else 12
    h1_threshold = all_sizes_sorted[0] if len(all_sizes_sorted) > 0 else 999
    h2_threshold = all_sizes_sorted[1] if len(all_sizes_sorted) > 1 else 999
    h3_threshold = all_sizes_sorted[2] if len(all_sizes_sorted) > 2 else 999

    md_lines: list[str] = []
    for page_idx, i in enumerate(pages):
        if page_idx > 0:
            md_lines.append("\n---\n")
        blocks = doc[i].get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]
        for block in blocks:
            if block["type"] == 1:
                continue  # image block — skip
            para_parts: list[str] = []
            heading_level = 0
            for line in block.get("lines", []):
                for span in line.get("spans", []):
                    text = span.get("text", "").strip()
                    if not text:
                        continue
                    size = span.get("size", body_size)
                    bold = bool(span.get("flags", 0) & (2 ** 4))

                    if size >= h1_threshold and size > body_size * 1.3:
                        heading_level = max(heading_level, 1)
                    elif size >= h2_threshold and size > body_size * 1.15:
                        heading_level = max(heading_level, 2)
                    elif size >= h3_threshold and size > body_size * 1.05:
                        heading_level = max(heading_level, 3)

                    if bold and heading_level == 0:
                        text = f"**{text}**"
                    para_parts.append(text)

            if not para_parts:
                continue
            combined = " ".join(para_parts)
            if heading_level == 1:
                md_lines.append(f"\n# {combined}\n")
            elif heading_level == 2:
                md_lines.append(f"\n## {combined}\n")
            elif heading_level == 3:
                md_lines.append(f"\n### {combined}\n")
            else:
                md_lines.append(combined)

    doc.close()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(md_lines), encoding="utf-8")
    return ConversionResult(output=output_path, pages=len(pages))


# ── pdf → docx ────────────────────────────────────────────────────────────

def _pdf_to_docx(
    input_path: Path,
    output_path: Path,
    *,
    page: Optional[int] = None,
    password: Optional[str] = None,
    **_options,
) -> ConversionResult:
    Converter = _require_pdf2docx()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    start = (page - 1) if page else None
    end   = page if page else None

    cv = Converter(str(input_path), password=password or "")
    try:
        cv.convert(str(output_path), start=start, end=end)
    finally:
        cv.close()

    return ConversionResult(output=output_path)


# ── pdf → csv ─────────────────────────────────────────────────────────────

def _pdf_to_csv(
    input_path: Path,
    output_path: Path,
    *,
    page: Optional[int] = None,
    password: Optional[str] = None,
    **_options,
) -> ConversionResult:
    import csv

    pdfplumber = _require_pdfplumber()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows_written = 0
    with pdfplumber.open(str(input_path), password=password or "") as pdf:
        pages_list = [pdf.pages[page - 1]] if page else pdf.pages
        with output_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for pg in pages_list:
                tables = pg.extract_tables()
                for table in tables:
                    for row in table:
                        writer.writerow([cell or "" for cell in row])
                    rows_written += len(table)

    if rows_written == 0:
        raise RuntimeError(
            "No tables found in the PDF. "
            "For plain text extraction use 'pdf → txt' instead."
        )
    return ConversionResult(output=output_path, rows=rows_written)


# ── pdf → xlsx ────────────────────────────────────────────────────────────

def _pdf_to_xlsx(
    input_path: Path,
    output_path: Path,
    *,
    page: Optional[int] = None,
    password: Optional[str] = None,
    **_options,
) -> ConversionResult:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    pdfplumber = _require_pdfplumber()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    total_rows = 0
    with pdfplumber.open(str(input_path), password=password or "") as pdf:
        pages_list = [pdf.pages[page - 1]] if page else pdf.pages
        for pg in pages_list:
            tables = pg.extract_tables()
            for tbl_idx, table in enumerate(tables):
                sheet_name = f"p{pg.page_number}_t{tbl_idx + 1}"
                ws = wb.create_sheet(title=sheet_name)
                for row in table:
                    ws.append([cell or "" for cell in row])
                total_rows += len(table)

    if not wb.sheetnames:
        raise RuntimeError(
            "No tables found in the PDF. "
            "For plain text extraction use 'pdf → txt' instead."
        )

    wb.save(str(output_path))
    return ConversionResult(output=output_path, rows=total_rows)


# ── pdf → json ────────────────────────────────────────────────────────────

def _pdf_to_json(
    input_path: Path,
    output_path: Path,
    *,
    page: Optional[int] = None,
    password: Optional[str] = None,
    **_options,
) -> ConversionResult:
    import json

    pdfplumber = _require_pdfplumber()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    result: list[dict] = []
    with pdfplumber.open(str(input_path), password=password or "") as pdf:
        pages_list = [pdf.pages[page - 1]] if page else pdf.pages
        for pg in pages_list:
            result.append({
                "page": pg.page_number,
                "text": pg.extract_text() or "",
                "tables": pg.extract_tables() or [],
            })

    output_path.write_text(
        __import__("json").dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return ConversionResult(output=output_path, pages=len(result))


# ── pdf → png / jpg ───────────────────────────────────────────────────────

def _pdf_to_image(
    input_path: Path,
    output_path: Path,
    fmt: str,
    *,
    page: Optional[int] = None,
    dpi: int = 150,
    password: Optional[str] = None,
    **_options,
) -> ConversionResult:
    """
    Renders PDF pages as images.
    - Single page (-p N): writes directly to output_path.
    - Multiple pages: writes a ZIP archive containing page_001.ext etc.
    """
    doc = _open_fitz(input_path, password)
    pages = _page_range(doc, page)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    pix_list: list[tuple[int, bytes]] = []
    for i in pages:
        matrix = doc[i].get_pixmap(dpi=dpi)
        img_bytes = matrix.tobytes(fmt)  # "png" or "jpeg"
        pix_list.append((i + 1, img_bytes))
    doc.close()

    if len(pix_list) == 1:
        output_path.write_bytes(pix_list[0][1])
        final_path = output_path
    else:
        # Multiple pages → zip so the user gets one output file
        zip_path = output_path.with_suffix(".zip")
        ext = "jpg" if fmt == "jpeg" else fmt
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for page_num, img_bytes in pix_list:
                zf.writestr(f"page_{page_num:03d}.{ext}", img_bytes)
        final_path = zip_path

    return ConversionResult(output=final_path, pages=len(pix_list))


def _pdf_to_png(input_path: Path, output_path: Path, **opts) -> ConversionResult:
    return _pdf_to_image(input_path, output_path, "png", **opts)


def _pdf_to_jpg(input_path: Path, output_path: Path, **opts) -> ConversionResult:
    return _pdf_to_image(input_path, output_path, "jpeg", **opts)


# ── register all routes ────────────────────────────────────────────────────

register(
    "pdf", "txt",
    backend="pymupdf",
    family="document",
    description="pdf → txt (text extraction; OCR fallback for scanned PDFs)",
    lossy=True,
    options=_TEXT_OPTIONS,
)(_pdf_to_txt)

register(
    "pdf", "html",
    backend="pymupdf",
    family="document",
    description="pdf → html (native layout export via pymupdf)",
    lossy=False,
    options=_TEXT_OPTIONS,
)(_pdf_to_html)

register(
    "pdf", "md",
    backend="pymupdf",
    family="document",
    description="pdf → md (text extraction with heading/bold inference)",
    lossy=True,
    options=_TEXT_OPTIONS,
)(_pdf_to_md)

register(
    "pdf", "docx",
    backend="pdf2docx",
    family="document",
    description="pdf → docx (layout-preserving, via pdf2docx)",
    lossy=False,
    options=_TEXT_OPTIONS,
)(_pdf_to_docx)

register(
    "pdf", "csv",
    backend="pdfplumber",
    family="document",
    description="pdf → csv (table extraction via pdfplumber)",
    lossy=False,
    options=_TABLE_OPTIONS,
)(_pdf_to_csv)

register(
    "pdf", "xlsx",
    backend="pdfplumber",
    family="document",
    description="pdf → xlsx (all tables, one sheet per table, via pdfplumber)",
    lossy=False,
    options=_TABLE_OPTIONS,
)(_pdf_to_xlsx)

register(
    "pdf", "json",
    backend="pdfplumber",
    family="document",
    description="pdf → json (structured: page text + tables, via pdfplumber)",
    lossy=False,
    options=_TABLE_OPTIONS,
)(_pdf_to_json)

register(
    "pdf", "png",
    backend="pymupdf",
    family="document",
    description="pdf → png (page render; multi-page output → zip of PNGs)",
    lossy=False,
    options=_IMAGE_OPTIONS,
)(_pdf_to_png)

register(
    "pdf", "jpg",
    backend="pymupdf",
    family="document",
    description="pdf → jpg (page render; multi-page output → zip of JPEGs)",
    lossy=True,
    options=_IMAGE_OPTIONS,
)(_pdf_to_jpg)
