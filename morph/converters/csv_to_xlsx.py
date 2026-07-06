"""
converters/csv_to_xlsx.py — csv -> xlsx, a fully self-contained converter.

Everything needed to turn a CSV into a polished, styled workbook lives in
this one file: encoding/delimiter sniffing, type inference, styling engine,
and the register() call with its full OptionSpec flag surface. No hidden
private engine module — what you see here is what runs.
"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Optional, cast

import chardet
import openpyxl
import openpyxl.cell
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..registry import ConversionResult, OptionSpec, register

# ─────────────────────────── helpers ────────────────────────────────────────

def _detect_encoding(path: Path) -> str:
    """Return best-guess encoding using chardet."""
    raw = path.read_bytes()
    result = chardet.detect(raw)
    enc = result.get("encoding") or "utf-8"
    if enc.lower() == "ascii":
        enc = "utf-8"
    return enc


def _detect_delimiter(path: Path, encoding: str) -> str:
    """Sniff the CSV delimiter; fall back to comma."""
    try:
        sample = path.read_text(encoding=encoding, errors="replace")[:8192]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def _infer_types(df: pd.DataFrame, date_input_fmt: Optional[str] = None) -> pd.DataFrame:
    """Coerce every column to the tightest appropriate type:
    bool -> bool, numeric -> float/int, datetime -> datetime, else str."""
    bool_map: dict[str, bool] = {
        "true": True, "false": False,
        "yes": True,  "no": False,
        "1": True,    "0": False,
    }

    for col in df.columns:
        series: pd.Series = cast(pd.Series, df[col])

        lower: pd.Series = cast(pd.Series, series.dropna()).astype(str).str.strip().str.lower()
        if len(lower) > 0 and lower.isin(bool_map).all():
            _bm = bool_map
            df[col] = series.map(
                lambda v, bm=_bm: bm.get(str(v).strip().lower(), v)  # type: ignore[return-value]
                if pd.notna(v) else v  # type: ignore[arg-type]
            )
            continue

        num: pd.Series = pd.to_numeric(series, errors="coerce")  # type: ignore[assignment]
        if num.notna().sum() / max(len(series), 1) >= 0.9:
            df[col] = num
            continue

        try:
            if date_input_fmt:
                dt: pd.Series = pd.to_datetime(  # type: ignore[assignment]
                    series, format=date_input_fmt, errors="coerce"
                )
            else:
                dt = pd.to_datetime(series, format="mixed", errors="coerce")  # type: ignore[assignment]
            if dt.notna().sum() / max(len(series), 1) >= 0.8:
                df[col] = dt
                continue
        except Exception:
            pass

    return df


def _px_width(text: str) -> float:
    return max(len(str(text)) + 2, 10)


def _apply_table_style(ws, n_rows: int, n_cols: int, table_style: str = "TableStyleMedium9") -> None:
    if n_rows < 1:
        return
    ref = f"A1:{get_column_letter(n_cols)}{n_rows + 1}"
    tbl = Table(displayName=f"Table_{ws.title.replace(' ', '_')}", ref=ref)
    style = TableStyleInfo(name=table_style, showFirstColumn=False, showLastColumn=False,
                            showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)


def _style_header(ws, n_cols: int, header_bg: str = "1F3864", header_fg: str = "FFFFFF") -> None:
    fill = PatternFill("solid", fgColor=header_bg)
    font = Font(bold=True, color=header_fg, size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(bottom=Side(style="medium", color="FFFFFF"))
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = border


def _apply_row_banding(ws, n_rows: int, n_cols: int, band_color: str = "EEF2FF") -> None:
    fill = PatternFill("solid", fgColor=band_color)
    for row_idx in range(3, n_rows + 2, 2):
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _autofit_columns(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_w = _px_width(col_name)
        for val in df[col_name]:
            max_w = max(max_w, _px_width(val))
        ws.column_dimensions[col_letter].width = min(max_w, 60)


def _apply_number_formats(ws, df: pd.DataFrame) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        dtype = df[col_name].dtype
        fmt = None
        if pd.api.types.is_datetime64_any_dtype(dtype):
            fmt = "YYYY-MM-DD HH:MM:SS"
        elif pd.api.types.is_float_dtype(dtype):
            fmt = "#,##0.00"
        elif pd.api.types.is_integer_dtype(dtype):
            fmt = "#,##0"
        if fmt:
            for row_idx in range(2, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = fmt


def _write_workbook(
    input_path: Path, output_path: Path, *,
    sheet_name: str = "Sheet1", delimiter: Optional[str] = None, encoding: Optional[str] = None,
    header: bool = True, infer_types: bool = True, date_input_fmt: Optional[str] = None,
    freeze_header: bool = True, freeze_cols: int = 0, add_table: bool = True,
    table_style: str = "TableStyleMedium9", row_banding: bool = True, band_color: str = "EEF2FF",
    header_bg: str = "1F3864", header_fg: str = "FFFFFF", autofit: bool = True,
    number_formats: bool = True, zoom: int = 100, row_height: Optional[float] = None,
    password: Optional[str] = None, workbook_password: Optional[str] = None,
    append_to_existing: bool = False, skiprows: int = 0, nrows: Optional[int] = None,
    usecols: Optional[str] = None,
) -> dict:
    """Convert a single CSV file to a styled Excel sheet. Returns {rows, cols, sheet, output}."""
    enc = encoding or _detect_encoding(input_path)
    delim = delimiter or _detect_delimiter(input_path, enc)

    read_kwargs: dict = dict(
        filepath_or_buffer=input_path, sep=delim, encoding=enc,
        header=0 if header else None, skiprows=skiprows if skiprows else None,
        nrows=nrows, dtype=str, keep_default_na=True, on_bad_lines="warn",
    )
    if usecols:
        cols = [c.strip() for c in usecols.split(",")]
        try:
            cols = [int(c) for c in cols]
        except ValueError:
            pass
        read_kwargs["usecols"] = cols

    df = pd.read_csv(**read_kwargs)

    if infer_types:
        df = _infer_types(df, date_input_fmt=date_input_fmt)

    n_rows, n_cols = df.shape

    if append_to_existing and output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        existing = wb.sheetnames
        base = sheet_name
        suffix = 1
        while sheet_name in existing:
            sheet_name = f"{base}_{suffix}"
            suffix += 1
        ws = wb.create_sheet(title=sheet_name)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = sheet_name

    if header:
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=str(col_name))

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = cast(openpyxl.cell.Cell, ws.cell(row=row_idx, column=col_idx))
            if pd.isna(value) if not isinstance(value, str) else False:  # type: ignore[arg-type]
                cell.value = None
            elif isinstance(value, (int, float, bool)):
                cell.value = value
            elif isinstance(value, datetime):
                cell.value = value
            else:
                cell.value = str(value) if value is not None else None

    if header and n_rows > 0:
        _style_header(ws, n_cols, header_bg=header_bg, header_fg=header_fg)
    if row_banding and n_rows > 0:
        _apply_row_banding(ws, n_rows, n_cols, band_color=band_color)
    if autofit:
        _autofit_columns(ws, df)
    if number_formats:
        _apply_number_formats(ws, df)

    if freeze_header or freeze_cols > 0:
        freeze_row = 2 if freeze_header and header else 1
        freeze_col = freeze_cols + 1
        ws.freeze_panes = f"{get_column_letter(freeze_col)}{freeze_row}"

    if add_table and header and n_rows > 0:
        _apply_table_style(ws, n_rows, n_cols, table_style=table_style)

    ws.row_dimensions[1].height = 20
    if row_height is not None:
        for row_idx in range(2, n_rows + 2):
            ws.row_dimensions[row_idx].height = row_height

    ws.sheet_view.zoomScale = max(10, min(zoom, 400))

    if password:
        ws.protection.sheet = True
        ws.protection.password = password
        ws.protection.enable()

    if workbook_password:
        wb.security.workbookPassword = workbook_password
        wb.security.lockStructure = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    return {"rows": n_rows, "cols": n_cols, "sheet": ws.title, "output": output_path}


# ─────────────────────────── registry wiring ────────────────────────────────

OPTIONS = [
    OptionSpec("delimiter", ("-d", "--delimiter"), "Field delimiter. Auto-detected if omitted."),
    OptionSpec("encoding", ("-e", "--encoding"), "Input encoding. Auto-detected if omitted."),
    OptionSpec("header", ("--no-header",), "Treat first row as data, not a header.", default=True, action="store_false"),
    OptionSpec("skiprows", ("--skip-rows",), "Rows to skip at the top of the CSV.", default=0, type=int),
    OptionSpec("nrows", ("--nrows",), "Maximum data rows to read.", type=int),
    OptionSpec("usecols", ("--cols",), "Comma-separated column names or indices to include."),
    OptionSpec("date_input_fmt", ("--date-fmt-in",), "strftime format of dates in the CSV."),
    OptionSpec("sheet_name", ("-s", "--sheet"), "Sheet name in the output workbook.", default="Sheet1"),
    OptionSpec("infer_types", ("--no-infer",), "Disable smart type inference.", default=True, action="store_false"),
    OptionSpec("freeze_header", ("--no-freeze",), "Disable frozen header row.", default=True, action="store_false"),
    OptionSpec("freeze_cols", ("--freeze-cols",), "Left-most columns to freeze.", default=0, type=int),
    OptionSpec("add_table", ("--no-table",), "Disable Excel Table / auto-filter.", default=True, action="store_false"),
    OptionSpec("table_style", ("--table-style",), "Excel table style name.", default="TableStyleMedium9"),
    OptionSpec("row_banding", ("--no-banding",), "Disable alternating row banding.", default=True, action="store_false"),
    OptionSpec("band_color", ("--band-color",), "Hex colour for banded rows (no #).", default="EEF2FF"),
    OptionSpec("autofit", ("--no-autofit",), "Disable auto column width.", default=True, action="store_false"),
    OptionSpec("number_formats", ("--no-formats",), "Disable number/date format strings.", default=True, action="store_false"),
    OptionSpec("header_bg", ("--header-bg",), "Header background hex colour (no #).", default="1F3864"),
    OptionSpec("header_fg", ("--header-fg",), "Header text hex colour (no #).", default="FFFFFF"),
    OptionSpec("zoom", ("--zoom",), "Sheet zoom level (10-400).", default=100, type=int),
    OptionSpec("row_height", ("--row-height",), "Data row height in points.", type=float),
    OptionSpec("password", ("--password",), "Password-protect the output sheet."),
    OptionSpec("workbook_password", ("--wb-password",), "Password-protect the workbook structure."),
    OptionSpec("append_to_existing", ("-a", "--append"), "Append as a new sheet into an existing workbook.",
               default=False, action="store_true"),
]

@register("csv", "xlsx", backend="native", family="data",
          description="csv → xlsx (styled workbook)", options=OPTIONS)
def csv_to_xlsx(input_path: Path, output_path: Path, **options) -> ConversionResult:
    stats = _write_workbook(input_path, output_path, **options)
    return ConversionResult(output=stats["output"], rows=stats["rows"])
