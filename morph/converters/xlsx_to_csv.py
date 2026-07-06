"""
converters/xlsx_to_csv.py — xlsx -> csv, a fully self-contained converter.

Sheet handling, merged-cell flattening, and the register() call with its
full OptionSpec flag surface all live in this one file.
"""

from __future__ import annotations

import csv
import warnings
from pathlib import Path
from typing import Optional, Union

import pandas as pd
from openpyxl import load_workbook

from ..registry import ConversionResult, OptionSpec, register

# ─────────────────────────── helpers ────────────────────────────────────────

def _get_sheet_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _get_visible_sheet_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=False, data_only=True)
    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    wb.close()
    return visible


def _flatten_merged_cells(path: Path, sheet: Union[str, int]) -> pd.DataFrame:
    """Read a sheet and fill every cell in a merged region with the
    top-left (anchor) value, then return as a DataFrame."""
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[sheet]

    merge_values: dict[tuple[int, int], object] = {}
    for merge_range in ws.merged_cells.ranges:
        anchor_val = ws.cell(merge_range.min_row, merge_range.min_col).value
        for r in range(merge_range.min_row, merge_range.max_row + 1):
            for c in range(merge_range.min_col, merge_range.max_col + 1):
                merge_values[(r, c)] = anchor_val

    rows: list[list] = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            key = (cell.row, cell.column)  # type: ignore[arg-type]
            row_data.append(merge_values[key] if key in merge_values else cell.value)
        rows.append(row_data)

    wb.close()
    if not rows:
        return pd.DataFrame()

    header = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(rows[0])]
    return pd.DataFrame(rows[1:], columns=header)


def _read_sheet(
    path: Path, sheet: Union[str, int], *, header_row: Optional[int], skiprows: int,
    skipfooter: int, usecols: Optional[list], nrows: Optional[int],
    strip_whitespace: bool, flatten_merges: bool,
) -> pd.DataFrame:
    if flatten_merges:
        df = _flatten_merged_cells(path, sheet)
        if strip_whitespace:
            for col in df.select_dtypes(include="object").columns:
                df[col] = df[col].str.strip()
        return df

    engine = "openpyxl"
    suffix = path.suffix.lower()
    if suffix == ".xls":
        engine = "xlrd"
    elif suffix == ".ods":
        engine = "odf"

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = pd.read_excel(
            path, sheet_name=sheet, header=header_row,
            skiprows=skiprows if skiprows else None, skipfooter=skipfooter,
            usecols=usecols, nrows=nrows, dtype=str, engine=engine, keep_default_na=True,
        )

    if strip_whitespace:
        for col in df.select_dtypes(include="object").columns:
            df[col] = df[col].str.strip()

    return df


def _df_to_csv(
    df: pd.DataFrame, output_path: Path, *, delimiter: str, encoding: str,
    line_terminator: str, quoting: int, na_rep: str, include_index: bool,
    date_format: Optional[str] = None,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(output_path, sep=delimiter, encoding=encoding, lineterminator=line_terminator,
              quoting=quoting, na_rep=na_rep, index=include_index,
              date_format=date_format or "%Y-%m-%d %H:%M:%S")


def _safe_filename(name: str) -> str:
    import re
    return re.sub(r'[\\/*?:"<>|]', "_", name)


def preview(
    input_path: Path, sheet: Union[str, int] = 0, *, nrows: int = 10,
    header_row: Optional[int] = 0, flatten_merges: bool = False, strip_whitespace: bool = True,
) -> pd.DataFrame:
    """First `nrows` rows of a sheet as a DataFrame — no file written."""
    return _read_sheet(input_path, sheet, header_row=header_row, skiprows=0, skipfooter=0,
                        usecols=None, nrows=nrows, strip_whitespace=strip_whitespace,
                        flatten_merges=flatten_merges)


def _convert_workbook(
    input_path: Path, output_path: Path, *, sheets: Optional[list] = None,
    skip_hidden: bool = True, all_sheets_separate: bool = False, combine_sheets: bool = False,
    combine_sheet_col: Optional[str] = "__sheet", delimiter: str = ",", encoding: str = "utf-8-sig",
    line_terminator: str = "\r\n", quoting: int = csv.QUOTE_MINIMAL, na_rep: str = "",
    date_format: Optional[str] = None, header_row: Optional[int] = 0, skiprows: int = 0,
    skipfooter: int = 0, usecols: Optional[str] = None, nrows: Optional[int] = None,
    strip_whitespace: bool = True, include_index: bool = False, flatten_merges: bool = False,
) -> list[dict]:
    """Convert an Excel workbook to one or more CSV files. Returns [{rows, cols, sheet, output}, ...]."""
    input_path = Path(input_path)
    output_path = Path(output_path)

    suffix = input_path.suffix.lower()
    if suffix == ".xls":
        all_sheet_names: list[str] = pd.ExcelFile(input_path, engine="xlrd").sheet_names
    elif suffix == ".ods":
        all_sheet_names = pd.ExcelFile(input_path, engine="odf").sheet_names
    else:
        all_sheet_names = _get_sheet_names(input_path)

    if skip_hidden and suffix not in (".xls", ".ods"):
        visible = _get_visible_sheet_names(input_path)
    else:
        visible = all_sheet_names

    if sheets:
        resolved: list[str] = []
        for s in sheets:
            if isinstance(s, int):
                if 0 <= s < len(all_sheet_names):
                    resolved.append(all_sheet_names[s])
            else:
                if s in all_sheet_names:
                    resolved.append(s)
        target_sheets = [s for s in resolved if s in visible]
    else:
        target_sheets = list(visible)

    if not target_sheets:
        raise ValueError("No matching visible sheets found in workbook.")

    parsed_usecols: Optional[list] = None
    if usecols:
        cols = [c.strip() for c in usecols.split(",")]
        try:
            parsed_usecols = [int(c) for c in cols]
        except ValueError:
            parsed_usecols = cols

    read_kw = dict(header_row=header_row, skiprows=skiprows, skipfooter=skipfooter,
                   usecols=parsed_usecols, nrows=nrows, strip_whitespace=strip_whitespace,
                   flatten_merges=flatten_merges)
    csv_kw = dict(delimiter=delimiter, encoding=encoding, line_terminator=line_terminator,
                  quoting=quoting, na_rep=na_rep, include_index=include_index, date_format=date_format)

    results: list[dict] = []

    if combine_sheets:
        frames: list[pd.DataFrame] = []
        for sheet in target_sheets:
            df = _read_sheet(input_path, sheet, **read_kw)  # type: ignore[arg-type]
            if combine_sheet_col:
                df.insert(0, combine_sheet_col, str(sheet))
            frames.append(df)
        combined = pd.concat(frames, ignore_index=True)
        _df_to_csv(combined, output_path, **csv_kw)  # type: ignore[arg-type]
        results.append({"rows": len(combined), "cols": len(combined.columns),
                         "sheet": f"{len(target_sheets)} sheets combined", "output": output_path})
        return results

    if all_sheets_separate or len(target_sheets) > 1:
        stem = output_path.stem
        parent = output_path.parent
        for sheet in target_sheets:
            safe = _safe_filename(str(sheet))
            out = parent / f"{stem}__{safe}.csv"
            df = _read_sheet(input_path, sheet, **read_kw)  # type: ignore[arg-type]
            _df_to_csv(df, out, **csv_kw)  # type: ignore[arg-type]
            results.append({"rows": len(df), "cols": len(df.columns), "sheet": sheet, "output": out})
        return results

    sheet = target_sheets[0]
    df = _read_sheet(input_path, sheet, **read_kw)  # type: ignore[arg-type]
    _df_to_csv(df, output_path, **csv_kw)  # type: ignore[arg-type]
    results.append({"rows": len(df), "cols": len(df.columns), "sheet": sheet, "output": output_path})
    return results


# ─────────────────────────── registry wiring ────────────────────────────────

OPTIONS = [
    OptionSpec("sheets", ("-s", "--sheets"), "Comma-separated sheet names/indices. Default: all visible."),
    OptionSpec("skip_hidden", ("--include-hidden",), "Include hidden sheets.", default=True, action="store_false"),
    OptionSpec("all_sheets_separate", ("--all-separate",), "One CSV per sheet even for a single match.",
               default=False, action="store_true"),
    OptionSpec("combine_sheets", ("--combine",), "Merge all sheets into one CSV.", default=False, action="store_true"),
    OptionSpec("combine_sheet_col", ("--combine-col",), "Sheet-identifier column name when combining.", default="__sheet"),
    OptionSpec("no_combine_col", ("--no-combine-col",), "Omit the sheet column when combining.",
               default=False, action="store_true"),
    OptionSpec("header_row", ("--header-row",), "0-based header row index. -1 for none.", default=0, type=int),
    OptionSpec("skiprows", ("--skip-rows",), "Rows to skip before the header.", default=0, type=int),
    OptionSpec("skipfooter", ("--skip-footer",), "Rows to skip at the bottom.", default=0, type=int),
    OptionSpec("nrows", ("--nrows",), "Maximum data rows to read per sheet.", type=int),
    OptionSpec("usecols", ("--cols",), "Comma-separated column names/indices to include."),
    OptionSpec("flatten_merges", ("--flatten-merges",), "Fill merged cells with the anchor value.",
               default=False, action="store_true"),
    OptionSpec("delimiter", ("-d", "--delimiter"), "Output field delimiter.", default=","),
    OptionSpec("encoding", ("-e", "--encoding"), "Output CSV encoding.", default="utf-8-sig"),
    OptionSpec("line_terminator", ("--line-term",), r"Line terminator (\r\n or \n).", default="\\r\\n"),
    OptionSpec("quoting", ("--quoting",), "minimal | all | nonnumeric | none", default="minimal"),
    OptionSpec("na_rep", ("--na",), "String to represent NaN / empty cells.", default=""),
    OptionSpec("date_format", ("--date-fmt",), "strftime format for date columns."),
    OptionSpec("strip_whitespace", ("--no-strip",), "Disable stripping whitespace from cells.",
               default=True, action="store_false"),
    OptionSpec("include_index", ("--index",), "Write row index as first column.", default=False, action="store_true"),
]

_QUOTING_MAP = {"minimal": csv.QUOTE_MINIMAL, "all": csv.QUOTE_ALL,
                "nonnumeric": csv.QUOTE_NONNUMERIC, "none": csv.QUOTE_NONE}


@register("xlsx", "csv", backend="native", family="data",
          description="xlsx → csv", lossy=True, options=OPTIONS)
def xlsx_to_csv(input_path: Path, output_path: Path, **options) -> ConversionResult:
    # small flag translations live right next to the flags they belong to
    if options.pop("no_combine_col", False):
        options["combine_sheet_col"] = None

    sheets = options.get("sheets")
    if sheets:
        parsed = []
        for p in [s.strip() for s in sheets.split(",")]:
            try:
                parsed.append(int(p))
            except ValueError:
                parsed.append(p)
        options["sheets"] = parsed

    header_row = options.get("header_row", 0)
    options["header_row"] = None if header_row == -1 else header_row

    options["line_terminator"] = options.get("line_terminator", "\\r\\n").replace("\\r\\n", "\r\n").replace("\\n", "\n")
    options["quoting"] = _QUOTING_MAP.get(str(options.get("quoting", "minimal")).lower(), csv.QUOTE_MINIMAL)

    results = _convert_workbook(input_path=input_path, output_path=output_path, **options)
    r = results[0]
    return ConversionResult(output=r["output"], rows=r["rows"])
